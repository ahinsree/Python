import pandas as pd
from datetime import datetime, timedelta

def create_python_learning_syllabus():
    # Calculate start date (next Monday)
    today = datetime.now()
    days_until_monday = (7 - today.weekday()) % 7
    if days_until_monday == 0:  # If today is Monday, start next Monday
        days_until_monday = 7
    start_date = today + timedelta(days=days_until_monday)
    start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    
    print(f"📅 Syllabus starts on: {start_date.strftime('%A, %B %d, %Y')}")
    
    # Create Main Syllabus Schedule data with dynamic dates
    main_syllabus_data = []
    
    current_date = start_date
    week_number = 1
    
    # Phase 1: Foundation (Months 1-2)
    foundation_weeks = [
        [1, 1, 'Programming basics, Python setup, IDEs', 'print(), variables, input/output', 'Hello World, Simple Calculator', 15],
        [1, 2, 'Data types, operators, basic operations', 'Strings, numbers, booleans, type conversion', 'Unit Converter, BMI Calculator', 18],
        [1, 3, 'Control flow, loops, boolean logic', 'if/else, for/while loops, range()', 'Number Guessing Game, Multiplication Table', 20],
        [1, 4, 'Functions, lists, tuples, strings', 'Function definition, list methods, slicing', 'Todo List Manager, Password Generator', 20],
        [2, 5, 'Dictionaries, sets, comprehensions', 'Dict operations, set theory, list comps', 'Student Gradebook, Word Counter', 18],
        [2, 6, 'File I/O, exception handling', 'Reading/writing files, try/except blocks', 'File Organizer, Log Analyzer', 20],
        [2, 7, 'OOP basics: classes, objects', 'Classes, inheritance, encapsulation', 'Bank Account System, Library Management', 22],
        [2, 8, 'Foundation Project Week', 'Review all basic concepts', 'Personal Finance Manager', 20]
    ]
    
    for month, week, topics, concepts, projects, hours in foundation_weeks:
        end_date = current_date + timedelta(days=6)
        main_syllabus_data.append([
            'Foundation', month, week, topics, concepts, projects, hours, 'Planned',
            current_date.strftime('%d-%b-%Y'), end_date.strftime('%d-%b-%Y'), ''
        ])
        current_date += timedelta(days=7)
        week_number += 1
    
    # Phase 2: Intermediate (Months 3-5)
    intermediate_weeks = [
        [3, 9, 'Web fundamentals, Flask intro', 'HTTP, REST APIs, Flask routing', 'Personal Blog, Weather App', 22],
        [3, 10, 'Database integration, SQL basics', 'SQLite, SQLAlchemy, ORM', 'User Registration, Product Catalog', 20],
        [3, 11, 'Frontend basics, API consumption', 'HTML/CSS, JavaScript, requests lib', 'Movie Database, News Aggregator', 25],
        [3, 12, 'Intermediate Project 1', 'Full-stack integration', 'E-commerce Store with Auth', 23],
        [4, 13, 'Data analysis with Pandas', 'DataFrames, cleaning, analysis', 'Sales Analysis, COVID Tracker', 22],
        [4, 14, 'Data visualization', 'Matplotlib, Seaborn, Plotly', 'Stock Visualizer, Survey Dashboard', 20],
        [4, 15, 'Web scraping, automation', 'BeautifulSoup, Selenium', 'Job Scraper, Social Media Bot', 25],
        [4, 16, 'Automation Project', 'Scripting and automation', 'Automated Reporting System', 18],
        [5, 17, 'Django framework basics', 'Django ORM, admin, views', 'Social Media Platform, CRM', 25],
        [5, 18, 'User authentication', 'JWT, OAuth, session management', 'Multi-user Blog, Admin Dashboard', 22],
        [5, 19, 'REST API development', 'DRF, API design, documentation', 'Todo API, E-commerce API', 23],
        [5, 20, 'Intermediate Capstone', 'Complete full-stack app', 'Project Management Tool', 30]
    ]
    
    for month, week, topics, concepts, projects, hours in intermediate_weeks:
        end_date = current_date + timedelta(days=6)
        main_syllabus_data.append([
            'Intermediate', month, week, topics, concepts, projects, hours, 'Planned',
            current_date.strftime('%d-%b-%Y'), end_date.strftime('%d-%b-%Y'), ''
        ])
        current_date += timedelta(days=7)
        week_number += 1
    
    # Phase 3: Advanced (Months 6-8)
    advanced_weeks = [
        [6, 21, 'Testing & TDD', 'unittest, pytest, mocking', 'Test suites for previous projects', 22],
        [6, 22, 'Code quality, debugging', 'PEP8, pylint, profiling', 'Code refactoring, optimization', 20],
        [6, 23, 'Design patterns', 'MVC, factory, singleton patterns', 'Refactored applications', 25],
        [6, 24, 'Advanced Project Struct', 'Enterprise architecture', 'Microservices application', 23],
        [7, 25, 'Database optimization', 'Indexing, query optimization', 'High-performance DB apps', 22],
        [7, 26, 'Caching & performance', 'Redis, Memcached, CDN', 'High-traffic web app', 20],
        [7, 27, 'Async programming', 'async/await, asyncio, FastAPI', 'Real-time Chat, WebSockets', 25],
        [7, 28, 'Performance Project', 'Scalability techniques', 'Real-time Analytics Dashboard', 28],
        [8, 29, 'Containerization', 'Docker, Docker Compose', 'Containerized applications', 22],
        [8, 30, 'Cloud deployment', 'AWS/GCP, CI/CD, serverless', 'Cloud-deployed apps', 25],
        [8, 31, 'Monitoring & logging', 'Logging practices, monitoring', 'Production-ready apps', 20],
        [8, 32, 'Deployment Project', 'Full deployment pipeline', 'SaaS Application', 30]
    ]
    
    for month, week, topics, concepts, projects, hours in advanced_weeks:
        end_date = current_date + timedelta(days=6)
        main_syllabus_data.append([
            'Advanced', month, week, topics, concepts, projects, hours, 'Planned',
            current_date.strftime('%d-%b-%Y'), end_date.strftime('%d-%b-%Y'), ''
        ])
        current_date += timedelta(days=7)
        week_number += 1
    
    # Phase 4: Specialization (Months 9-12)
    specialization_weeks = [
        [9, 33, 'Track A: Advanced Pandas/NumPy', 'Data manipulation, analysis', 'Data analysis projects', 22],
        [9, 34, 'Track A: Advanced Pandas/NumPy', 'Data cleaning, transformation', 'Data processing projects', 22],
        [9, 35, 'Track A: ML with Scikit-learn', 'Machine learning basics', 'Predictive models', 25],
        [9, 36, 'Track A: ML with Scikit-learn', 'Model evaluation, tuning', 'ML pipeline projects', 25],
        [10, 37, 'Track A: Deep Learning', 'Neural networks, TensorFlow', 'Deep learning projects', 25],
        [10, 38, 'Track A: Deep Learning', 'CNN, RNN, transfer learning', 'Advanced ML projects', 25],
        [10, 39, 'Track A: Data Engineering', 'Big data, pipelines', 'Data engineering projects', 28],
        [10, 40, 'Track A: Data Engineering', 'ETL processes, data lakes', 'Scalable data systems', 28],
        [11, 41, 'Major Portfolio Project', 'Project planning, setup', 'Portfolio project phase 1', 25],
        [11, 42, 'Major Portfolio Project', 'Development, implementation', 'Portfolio project phase 2', 25],
        [11, 43, 'Major Portfolio Project', 'Testing, refinement', 'Portfolio project phase 3', 25],
        [11, 44, 'Major Portfolio Project', 'Deployment, documentation', 'Portfolio project completion', 25],
        [12, 45, 'Open source contribution', 'Finding projects, contributing', 'OS contributions', 20],
        [12, 46, 'Open source contribution', 'PR reviews, collaboration', 'Meaningful contributions', 20],
        [12, 47, 'Interview preparation', 'LeetCode, system design', 'Mock interviews', 22],
        [12, 48, 'Portfolio & job search', 'Resume, networking, applications', 'Job ready', 18]
    ]
    
    for month, week, topics, concepts, projects, hours in specialization_weeks:
        end_date = current_date + timedelta(days=6)
        main_syllabus_data.append([
            'Specialization', month, week, topics, concepts, projects, hours, 'Planned',
            current_date.strftime('%d-%b-%Y'), end_date.strftime('%d-%b-%Y'), ''
        ])
        current_date += timedelta(days=7)
        week_number += 1
    
    # Calculate completion date
    completion_date = current_date - timedelta(days=7)
    print(f"🎯 Projected completion date: {completion_date.strftime('%A, %B %d, %Y')}")
    print(f"📅 Total duration: 48 weeks (~11 months)")
    
    # Create Progress Tracker data
    progress_data = []
    for i, row in enumerate(main_syllabus_data):
        progress_data.append([
            i + 1,  # Week
            row[1],  # Month
            row[0],  # Phase
            row[3],  # Topics
            row[6],  # Hours Planned
            0,       # Hours Completed
            '0%',    # Completion %
            'Not Started',  # Status
            'No',    # Project Completed
            '',      # Notes
            ''       # Date Completed
        ])
    
    # Create Project Portfolio data with updated months/weeks
    project_data = [
        ['Hello World', 'Mini-project', 'Beginner', 1, 1, 'Planned', '', 'Python, basics', 'First Python program', ''],
        ['Simple Calculator', 'Mini-project', 'Beginner', 1, 1, 'Planned', '', 'Python, functions', 'Basic arithmetic operations', ''],
        ['Unit Converter', 'Mini-project', 'Beginner', 1, 2, 'Planned', '', 'Python, math ops', 'Convert between units', ''],
        ['BMI Calculator', 'Mini-project', 'Beginner', 1, 2, 'Planned', '', 'Python, conditionals', 'Health metric calculator', ''],
        ['Number Guessing Game', 'Mini-project', 'Beginner', 1, 3, 'Planned', '', 'Python, loops', 'Interactive game', ''],
        ['Multiplication Table', 'Mini-project', 'Beginner', 1, 3, 'Planned', '', 'Python, loops', 'Educational tool', ''],
        ['Todo List Manager', 'App', 'Intermediate', 1, 4, 'Planned', '', 'Python, file I/O', 'Task management', ''],
        ['Password Generator', 'App', 'Intermediate', 1, 4, 'Planned', '', 'Python, random', 'Security tool', ''],
        ['Student Gradebook', 'App', 'Intermediate', 2, 5, 'Planned', '', 'Python, dictionaries', 'Academic tracking', ''],
        ['Word Frequency Counter', 'App', 'Intermediate', 2, 5, 'Planned', '', 'Python, text processing', 'Text analysis', ''],
        ['File Organizer', 'Utility', 'Intermediate', 2, 6, 'Planned', '', 'Python, OS module', 'File management', ''],
        ['Log Analyzer', 'Utility', 'Intermediate', 2, 6, 'Planned', '', 'Python, regex', 'System monitoring', ''],
        ['Bank Account System', 'App', 'Advanced', 2, 7, 'Planned', '', 'Python, OOP', 'Financial simulation', ''],
        ['Library Management', 'App', 'Advanced', 2, 7, 'Planned', '', 'Python, OOP', 'Inventory system', ''],
        ['Personal Finance Manager', 'Capstone', 'Advanced', 2, 8, 'Planned', '', 'Python, full stack', 'Comprehensive finance app', '']
    ]
    
    # Create Skills Checklist data
    skills_data = [
        ['Python Basics', 'Variables and Data Types', 1, 'Not Started', '0%', '', 'Python docs'],
        ['Python Basics', 'Control Structures', 1, 'Not Started', '0%', '', 'Online tutorials'],
        ['Python Basics', 'Functions', 1, 'Not Started', '0%', '', 'Practice exercises'],
        ['Data Structures', 'Lists and Tuples', 1, 'Not Started', '0%', '', 'Coding challenges'],
        ['Data Structures', 'Dictionaries and Sets', 2, 'Not Started', '0%', '', 'Real projects'],
        ['File Handling', 'Reading/Writing Files', 2, 'Not Started', '0%', '', 'File I/O exercises'],
        ['OOP', 'Classes and Objects', 2, 'Not Started', '0%', '', 'OOP tutorials'],
        ['Web Development', 'Flask Basics', 3, 'Not Started', '0%', '', 'Flask documentation'],
        ['Web Development', 'HTML/CSS Integration', 3, 'Not Started', '0%', '', 'Web tutorials'],
        ['Databases', 'SQL Basics', 3, 'Not Started', '0%', '', 'SQL practice'],
        ['Databases', 'ORM with SQLAlchemy', 3, 'Not Started', '0%', '', 'Database projects'],
        ['APIs', 'REST API Concepts', 3, 'Not Started', '0%', '', 'API tutorials'],
        ['Data Analysis', 'Pandas DataFrames', 4, 'Not Started', '0%', '', 'Data analysis courses'],
        ['Data Visualization', 'Matplotlib/Seaborn', 4, 'Not Started', '0%', '', 'Visualization guides'],
        ['Web Scraping', 'BeautifulSoup', 4, 'Not Started', '0%', '', 'Scraping projects'],
        ['Web Framework', 'Django Basics', 5, 'Not Started', '0%', '', 'Django documentation'],
        ['Authentication', 'User Auth Systems', 5, 'Not Started', '0%', '', 'Security tutorials'],
        ['API Development', 'Django REST Framework', 5, 'Not Started', '0%', '', 'API development'],
        ['Testing', 'Unit Testing', 6, 'Not Started', '0%', '', 'Testing frameworks'],
        ['Code Quality', 'PEP8, Linting', 6, 'Not Started', '0%', '', 'Style guides']
    ]
    
    # Create Weekly Schedule data with current year times
    current_year = datetime.now().year
    schedule_data = [
        ['Monday', '18:00-19:00', 'Theory Learning', 'New Concepts', '1 hour', 'No', ''],
        ['Monday', '19:00-21:00', 'Practice', 'Exercises', '2 hours', 'No', ''],
        ['Tuesday', '18:00-20:00', 'Coding Practice', 'Challenges', '2 hours', 'No', ''],
        ['Tuesday', '20:00-22:00', 'Project Work', 'Current Project', '2 hours', 'No', ''],
        ['Wednesday', '18:00-22:00', 'Project Development', 'Main Project', '4 hours', 'No', ''],
        ['Thursday', '18:00-19:00', 'Review', 'Previous Topics', '1 hour', 'No', ''],
        ['Thursday', '19:00-21:00', 'Debugging', 'Problem Solving', '2 hours', 'No', ''],
        ['Friday', '18:00-21:00', 'Advanced Topics', 'Deep Dive', '3 hours', 'No', ''],
        ['Saturday', '10:00-16:00', 'Project Marathon', 'Extended Work', '6 hours', 'No', ''],
        ['Sunday', '16:00-18:00', 'Planning', 'Next Week', '2 hours', 'No', '']
    ]
    
    # Create DataFrames
    df_main = pd.DataFrame(main_syllabus_data, 
                          columns=['Phase', 'Month', 'Week', 'Topics', 'Key Concepts', 
                                  'Projects', 'Hours', 'Status', 'Start Date', 'End Date', 'Notes'])
    
    df_progress = pd.DataFrame(progress_data,
                              columns=['Week', 'Month', 'Phase', 'Topics', 'Hours Planned',
                                      'Hours Completed', 'Completion %', 'Status', 
                                      'Project Completed', 'Notes', 'Date Completed'])
    
    df_projects = pd.DataFrame(project_data,
                              columns=['Project Name', 'Type', 'Difficulty', 'Month', 'Week',
                                      'Status', 'GitHub Link', 'Technologies Used', 
                                      'Description', 'Demo Link'])
    
    df_skills = pd.DataFrame(skills_data,
                            columns=['Skill Category', 'Specific Skill', 'Month Target',
                                    'Status', 'Confidence Level', 'Notes', 'Resources'])
    
    df_schedule = pd.DataFrame(schedule_data,
                              columns=['Day', 'Time Slot', 'Activity Type', 'Topic',
                                      'Duration', 'Completed', 'Notes'])
    
    # Create Excel file with multiple sheets
    filename = f'Python_Learning_Syllabus_{start_date.strftime("%Y%m%d")}.xlsx'
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Main Syllabus Schedule
        df_main.to_excel(writer, sheet_name='MAIN SYLLABUS SCHEDULE', index=False)
        
        # Progress Tracker
        df_progress.to_excel(writer, sheet_name='PROGRESS TRACKER', index=False)
        
        # Project Portfolio
        df_projects.to_excel(writer, sheet_name='PROJECT PORTFOLIO', index=False)
        
        # Skills Checklist
        df_skills.to_excel(writer, sheet_name='SKILLS CHECKLIST', index=False)
        
        # Weekly Schedule Template
        df_schedule.to_excel(writer, sheet_name='WEEKLY SCHEDULE', index=False)
        
        # Get workbook and worksheets for formatting
        workbook = writer.book
        
        # Adjust column widths for better readability
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✅ Excel file '{filename}' created successfully!")
    print("📊 File contains 5 sheets:")
    print("   1. MAIN SYLLABUS SCHEDULE")
    print("   2. PROGRESS TRACKER") 
    print("   3. PROJECT PORTFOLIO")
    print("   4. SKILLS CHECKLIST")
    print("   5. WEEKLY SCHEDULE")
    print(f"\n📁 File saved in your current directory")
    
    return filename

def add_excel_formatting(filename):
    """Add advanced formatting to the Excel file"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl import load_workbook
        
        # Load the workbook
        workbook = load_workbook(filename)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        phase_fills = {
            'Foundation': PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
            'Intermediate': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            'Advanced': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            'Specialization': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        }
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Apply formatting to all sheets
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Format header row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            # Apply phase-based coloring to MAIN SYLLABUS sheet
            if sheet_name == 'MAIN SYLLABUS SCHEDULE':
                for row in range(2, worksheet.max_row + 1):
                    phase_cell = worksheet[f'A{row}']
                    if phase_cell.value in phase_fills:
                        for col in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row, column=col).fill = phase_fills[phase_cell.value]
                            worksheet.cell(row=row, column=col).border = thin_border
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
            # Auto-filter on header row
            worksheet.auto_filter.ref = worksheet.dimensions
        
        # Save the formatted workbook
        formatted_filename = filename.replace('.xlsx', '_FORMATTED.xlsx')
        workbook.save(formatted_filename)
        print(f"✅ Formatted version saved as '{formatted_filename}'")
        
    except ImportError:
        print("ℹ️  For advanced formatting, install openpyxl: pip install openpyxl")

def display_timeline():
    """Display a simple timeline of the learning journey"""
    today = datetime.now()
    days_until_monday = (7 - today.weekday()) % 7
    if days_until_monday == 0:
        days_until_monday = 7
    start_date = today + timedelta(days=days_until_monday)
    
    milestones = [
        (1, "Python Basics", start_date + timedelta(weeks=3)),
        (2, "Foundation Complete", start_date + timedelta(weeks=7)),
        (3, "Web Development", start_date + timedelta(weeks=11)),
        (4, "Intermediate Complete", start_date + timedelta(weeks=19)),
        (5, "Advanced Topics", start_date + timedelta(weeks=27)),
        (6, "Specialization", start_date + timedelta(weeks=35)),
        (7, "Portfolio Ready", start_date + timedelta(weeks=47))
    ]
    
    print("\n📅 LEARNING TIMELINE:")
    print("=" * 50)
    for milestone, description, date in milestones:
        print(f"Month {milestone:2d}: {description:25} → {date.strftime('%b %d, %Y')}")

if __name__ == "__main__":
    print("🐍 Python Learning Syllabus Generator")
    print("=" * 40)
    
    # Create the main Excel file
    filename = create_python_learning_syllabus()
    
    # Display timeline
    display_timeline()
    
    # Ask if user wants formatted version
    try:
        format_choice = input("\nWould you like a formatted version with colors? (y/n): ").lower()
        if format_choice in ['y', 'yes']:
            add_excel_formatting(filename)
    except:
        print("ℹ️  Basic version created successfully.")
    
    print("\n🎉 Your personalized Python Learning Syllabus is ready!")
    print("💡 You can now open the Excel file and start tracking your progress")
    print("🚀 Happy learning! 🐍")