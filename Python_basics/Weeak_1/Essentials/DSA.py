import pandas as pd
from datetime import datetime, timedelta

def get_start_date_from_user():
    """Interactively gets the desired start date from the user."""
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    while True:
        print("\nSelect a start date option:")
        print("  1. Start Today")
        print("  2. Start Next Monday")
        print("  3. Enter a custom date (YYYY-MM-DD)")
        choice = input("Enter your choice (1, 2, or 3): ")

        if choice == '1':
            return today
        elif choice == '2':
            days_until_monday = (7 - today.weekday()) % 7
            if days_until_monday == 0:  # If today is Monday, start next Monday
                days_until_monday = 7
            return today + timedelta(days=days_until_monday)
        elif choice == '3':
            while True:
                date_str = input("Enter the start date (YYYY-MM-DD): ")
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except ValueError:
                    print("❌ Invalid format. Please use YYYY-MM-DD.")
        else:
            print("❌ Invalid choice. Please enter 1, 2, or 3.")

def create_dsa_learning_syllabus():
    """Generates the main DSA learning syllabus Excel file."""
    # Get start date from user for better convenience
    start_date = get_start_date_from_user()
    
    print(f"📅 DSA Syllabus starts on: {start_date.strftime('%A, %B %d, %Y')}")
    
    # Create Main Syllabus Schedule data with dynamic dates
    main_syllabus_data = []
    
    current_date = start_date
    week_number = 1
    
    # Phase 1: Foundation & Complexity Analysis (Weeks 1-4)
    foundation_weeks = [
        [1, 1, 'Big O Notation & Complexity Analysis', 'Time/Space complexity, asymptotic analysis', 'Complexity analysis of basic operations', 12],
        [1, 2, 'Arrays, Lists & Basic Operations', 'Array operations, memory allocation, Python lists', 'Array manipulator, basic algorithms', 15],
        [1, 3, 'Strings & Character Manipulation', 'String operations, pattern matching, encoding', 'String reverser, palindrome checker', 14],
        [1, 4, 'Hash Tables & Dictionaries', 'Hashing, collision resolution, Python dicts', 'Word frequency counter, cache implementation', 16]
    ]
    
    for month, week, topics, concepts, projects, hours in foundation_weeks:
        end_date = current_date + timedelta(days=6)
        main_syllabus_data.append([
            'Foundation', month, week, topics, concepts, projects, hours, 'Planned',
            current_date.strftime('%d-%b-%Y'), end_date.strftime('%d-%b-%Y'), ''
        ])
        current_date += timedelta(days=7)
        week_number += 1
    
    # Phase 2: Linear Data Structures (Weeks 5-8)
    linear_ds_weeks = [
        [2, 5, 'Linked Lists - Singly & Doubly', 'Node structure, pointers, traversal', 'Linked list implementation, music playlist', 18],
        [2, 6, 'Stacks & Applications', 'LIFO principle, stack operations', 'Expression evaluator, browser history', 16],
        [2, 7, 'Queues & Variations', 'FIFO principle, circular queues, deques', 'Task scheduler, print queue simulator', 17],
        [2, 8, 'Linear Structures Project Week', 'Combining lists, stacks, queues', 'Text editor with undo/redo', 20]
    ]
    
    for month, week, topics, concepts, projects, hours in linear_ds_weeks:
        end_date = current_date + timedelta(days=6)
        main_syllabus_data.append([
            'Linear Structures', month, week, topics, concepts, projects, hours, 'Planned',
            current_date.strftime('%d-%b-%Y'), end_date.strftime('%d-%b-%Y'), ''
        ])
        current_date += timedelta(days=7)
        week_number += 1
    
    # Phase 3: Trees & Hierarchical Data (Weeks 9-12)
    trees_weeks = [
        [3, 9, 'Binary Trees & Traversals', 'Tree terminology, DFS traversals', 'File system navigator, expression trees', 20],
        [3, 10, 'Binary Search Trees', 'BST properties, search/insert/delete', 'Dictionary implementation, auto-complete', 18],
        [3, 11, 'Balanced Trees (AVL/Red-Black)', 'Tree rotations, balancing algorithms', 'Balanced BST implementation', 22],
        [3, 12, 'Tree Applications & Heaps', 'Heap operations, priority queues', 'Task scheduler with priorities', 19]
    ]
    
    for month, week, topics, concepts, projects, hours in trees_weeks:
        end_date = current_date + timedelta(days=6)
        main_syllabus_data.append([
            'Trees & Hierarchical', month, week, topics, concepts, projects, hours, 'Planned',
            current_date.strftime('%d-%b-%Y'), end_date.strftime('%d-%b-%Y'), ''
        ])
        current_date += timedelta(days=7)
        week_number += 1
    
    # Phase 4: Graphs & Algorithms (Weeks 13-16)
    graphs_weeks = [
        [4, 13, 'Graph Fundamentals & Representations', 'Adjacency list/matrix, graph types', 'Social network graph, maze generator', 21],
        [4, 14, 'Graph Traversal Algorithms', 'BFS, DFS, connected components', 'Web crawler, friend recommendation', 20],
        [4, 15, 'Shortest Path Algorithms', 'Dijkstra, Bellman-Ford, A* search', 'GPS navigation, network routing', 23],
        [4, 16, 'Graph Applications & MST', 'Minimum spanning trees, topological sort', 'Network design, course scheduler', 22]
    ]
    
    for month, week, topics, concepts, projects, hours in graphs_weeks:
        end_date = current_date + timedelta(days=6)
        main_syllabus_data.append([
            'Graphs & Algorithms', month, week, topics, concepts, projects, hours, 'Planned',
            current_date.strftime('%d-%b-%Y'), end_date.strftime('%d-%b-%Y'), ''
        ])
        current_date += timedelta(days=7)
        week_number += 1
    
    # Phase 5: Sorting & Searching (Weeks 17-20)
    sorting_weeks = [
        [5, 17, 'Basic Sorting Algorithms', 'Bubble, Selection, Insertion sort', 'Sorting visualizer, performance comparison', 18],
        [5, 18, 'Advanced Sorting Algorithms', 'Merge sort, Quick sort, Heap sort', 'Large dataset sorter, hybrid algorithms', 20],
        [5, 19, 'Searching Algorithms', 'Binary search, interpolation search', 'Database query optimizer, search engine', 17],
        [5, 20, 'Non-comparison Sorts', 'Counting sort, Radix sort, Bucket sort', 'Specialized sorting for specific data', 19]
    ]
    
    for month, week, topics, concepts, projects, hours in sorting_weeks:
        end_date = current_date + timedelta(days=6)
        main_syllabus_data.append([
            'Sorting & Searching', month, week, topics, concepts, projects, hours, 'Planned',
            current_date.strftime('%d-%b-%Y'), end_date.strftime('%d-%b-%Y'), ''
        ])
        current_date += timedelta(days=7)
        week_number += 1
    
    # Phase 6: Dynamic Programming (Weeks 21-24)
    dp_weeks = [
        [6, 21, 'DP Fundamentals & Memoization', 'Overlapping subproblems, optimal substructure', 'Fibonacci optimizer, memoization wrapper', 20],
        [6, 22, '1D Dynamic Programming', 'Knapsack variations, sequence problems', 'Stock trader, rod cutting optimizer', 22],
        [6, 23, '2D Dynamic Programming', 'Matrix problems, string alignment', 'Text diff tool, path finding', 23],
        [6, 24, 'Advanced DP Patterns', 'State machine DP, bitmask DP', 'Game strategy optimizer, TSP solver', 21]
    ]
    
    for month, week, topics, concepts, projects, hours in dp_weeks:
        end_date = current_date + timedelta(days=6)
        main_syllabus_data.append([
            'Dynamic Programming', month, week, topics, concepts, projects, hours, 'Planned',
            current_date.strftime('%d-%b-%Y'), end_date.strftime('%d-%b-%Y'), ''
        ])
        current_date += timedelta(days=7)
        week_number += 1
    
    # Phase 7: Advanced Topics (Weeks 25-28)
    advanced_weeks = [
        [7, 25, 'Greedy Algorithms', 'Greedy choice property, matroids', 'Huffman coding, interval scheduling', 19],
        [7, 26, 'Backtracking & Recursion', 'State space search, pruning', 'Sudoku solver, N-queens problem', 20],
        [7, 27, 'Tries & String Algorithms', 'Trie operations, pattern matching', 'Autocomplete system, spell checker', 18],
        [7, 28, 'Advanced Data Structures', 'Segment trees, Fenwick trees', 'Range query processor, inventory system', 22]
    ]
    
    for month, week, topics, concepts, projects, hours in advanced_weeks:
        end_date = current_date + timedelta(days=6)
        main_syllabus_data.append([
            'Advanced Topics', month, week, topics, concepts, projects, hours, 'Planned',
            current_date.strftime('%d-%b-%Y'), end_date.strftime('%d-%b-%Y'), ''
        ])
        current_date += timedelta(days=7)
        week_number += 1
    
    # Phase 8: Interview Preparation (Weeks 29-32)
    interview_weeks = [
        [8, 29, 'Problem Solving Patterns', 'Two pointers, sliding window, fast/slow', 'Pattern recognition exercises', 20],
        [8, 30, 'LeetCode Strategy & Techniques', 'Problem categorization, time management', 'Mock interview simulations', 22],
        [8, 31, 'System Design Basics', 'Scalability, database design, APIs', 'URL shortener, chat system design', 21],
        [8, 32, 'Final Review & Portfolio Prep', 'Comprehensive review, project polish', 'Portfolio website, GitHub optimization', 18]
    ]
    
    for month, week, topics, concepts, projects, hours in interview_weeks:
        end_date = current_date + timedelta(days=6)
        main_syllabus_data.append([
            'Interview Prep', month, week, topics, concepts, projects, hours, 'Planned',
            current_date.strftime('%d-%b-%Y'), end_date.strftime('%d-%b-%Y'), ''
        ])
        current_date += timedelta(days=7)
        week_number += 1
    
    # Calculate completion date
    completion_date = current_date - timedelta(days=7)
    print(f"🎯 Projected completion date: {completion_date.strftime('%A, %B %d, %Y')}")
    print(f"📅 Total duration: 32 weeks (~8 months)")
    
    # Create Progress Tracker data
    progress_data = []
    for i, row in enumerate(main_syllabus_data):
        progress_data.append([
            i + 1,  # Week
            row[1],  # Month
            row[0],  # Phase
            row[3],  # Topics
            row[6],  # Hours Planned
            0,       # Hours Completed
            '0%',    # Completion %
            'Not Started',  # Status
            'No',    # Project Completed
            '',      # Notes
            ''       # Date Completed
        ])
    
    # Create Project Portfolio data
    project_data = [
        ['Complexity Analyzer', 'Tool', 'Beginner', 1, 1, 'Planned', '', 'Python, timeit', 'Algorithm performance analysis', ''],
        ['Array Manipulator', 'Library', 'Beginner', 1, 2, 'Planned', '', 'Python, arrays', 'Custom array operations', ''],
        ['String Processor', 'Utility', 'Beginner', 1, 3, 'Planned', '', 'Python, strings', 'Text processing utilities', ''],
        ['Hash Map Implementation', 'Library', 'Intermediate', 1, 4, 'Planned', '', 'Python, hashing', 'Custom dictionary implementation', ''],
        ['Music Playlist Manager', 'Application', 'Intermediate', 2, 5, 'Planned', '', 'Python, linked lists', 'Playlist with navigation', ''],
        ['Expression Evaluator', 'Tool', 'Intermediate', 2, 6, 'Planned', '', 'Python, stacks', 'Mathematical expression parser', ''],
        ['Task Scheduler', 'Application', 'Intermediate', 2, 7, 'Planned', '', 'Python, queues', 'Priority-based task management', ''],
        ['Text Editor with Undo', 'Application', 'Advanced', 2, 8, 'Planned', '', 'Python, stacks', 'Full-featured text editor', ''],
        ['File System Navigator', 'Utility', 'Intermediate', 3, 9, 'Planned', '', 'Python, trees', 'Directory tree visualization', ''],
        ['Auto-complete System', 'Application', 'Advanced', 3, 10, 'Planned', '', 'Python, BST', 'Word prediction engine', ''],
        ['Balanced Dictionary', 'Library', 'Advanced', 3, 11, 'Planned', '', 'Python, AVL trees', 'Self-balancing dictionary', ''],
        ['Priority Task Manager', 'Application', 'Intermediate', 3, 12, 'Planned', '', 'Python, heaps', 'Task management with priorities', ''],
        ['Social Network Analyzer', 'Application', 'Advanced', 4, 13, 'Planned', '', 'Python, graphs', 'Friend recommendation system', ''],
        ['Maze Solver', 'Game', 'Intermediate', 4, 14, 'Planned', '', 'Python, BFS/DFS', 'Pathfinding visualization', ''],
        ['GPS Navigation System', 'Application', 'Advanced', 4, 15, 'Planned', '', 'Python, Dijkstra', 'Shortest path finder', ''],
        ['Course Scheduler', 'Tool', 'Advanced', 4, 16, 'Planned', '', 'Python, topological sort', 'Academic schedule planner', '']
    ]
    
    # Create Skills Checklist data
    skills_data = [
        ['Complexity Analysis', 'Big O Notation', 1, 'Not Started', '0%', '', 'Algorithm analysis books'],
        ['Arrays & Lists', 'Python list operations', 1, 'Not Started', '0%', '', 'Python documentation'],
        ['Strings', 'String manipulation', 1, 'Not Started', '0%', '', 'String algorithm tutorials'],
        ['Hash Tables', 'Dictionary implementation', 1, 'Not Started', '0%', '', 'Hashing tutorials'],
        ['Linked Lists', 'Singly/Doubly linked', 2, 'Not Started', '0%', '', 'Data structure courses'],
        ['Stacks', 'LIFO operations', 2, 'Not Started', '0%', '', 'Stack applications'],
        ['Queues', 'FIFO operations', 2, 'Not Started', '0%', '', 'Queue variations'],
        ['Trees', 'Binary tree traversals', 3, 'Not Started', '0%', '', 'Tree algorithms'],
        ['BST', 'Search/Insert/Delete', 3, 'Not Started', '0%', '', 'BST tutorials'],
        ['Balanced Trees', 'AVL rotations', 3, 'Not Started', '0%', '', 'Balancing algorithms'],
        ['Heaps', 'Priority queue operations', 3, 'Not Started', '0%', '', 'Heap applications'],
        ['Graphs', 'BFS/DFS algorithms', 4, 'Not Started', '0%', '', 'Graph theory'],
        ['Shortest Path', 'Dijkstra algorithm', 4, 'Not Started', '0%', '', 'Pathfinding guides'],
        ['Sorting', 'Quick/Merge sort', 5, 'Not Started', '0%', '', 'Sorting visualizations'],
        ['Searching', 'Binary search', 5, 'Not Started', '0%', '', 'Search algorithms'],
        ['Dynamic Programming', 'Memoization', 6, 'Not Started', '0%', '', 'DP patterns'],
        ['Greedy Algorithms', 'Optimal choices', 7, 'Not Started', '0%', '', 'Greedy strategies'],
        ['Backtracking', 'State space search', 7, 'Not Started', '0%', '', 'Backtracking tutorials'],
        ['Tries', 'Prefix trees', 7, 'Not Started', '0%', '', 'Trie applications'],
        ['Interview Prep', 'Problem solving', 8, 'Not Started', '0%', '', 'LeetCode patterns']
    ]
    
    # Create Weekly Schedule data
    schedule_data = [
        ['Monday', '18:00-19:30', 'Theory Learning', 'New Concepts & Algorithms', '1.5 hours', 'No', ''],
        ['Monday', '19:30-21:00', 'Code Implementation', 'Data Structure Implementation', '1.5 hours', 'No', ''],
        ['Tuesday', '18:00-20:00', 'Algorithm Practice', 'Problem Solving', '2 hours', 'No', ''],
        ['Tuesday', '20:00-22:00', 'Project Work', 'Weekly Project', '2 hours', 'No', ''],
        ['Wednesday', '18:00-21:00', 'LeetCode/Exercises', 'Pattern Recognition', '3 hours', 'No', ''],
        ['Thursday', '18:00-19:00', 'Review Session', 'Previous Topics', '1 hour', 'No', ''],
        ['Thursday', '19:00-21:00', 'Debugging & Optimization', 'Code Improvement', '2 hours', 'No', ''],
        ['Friday', '18:00-20:00', 'Advanced Topics', 'Deep Dive', '2 hours', 'No', ''],
        ['Saturday', '10:00-16:00', 'Project Marathon', 'Extended Implementation', '6 hours', 'No', ''],
        ['Sunday', '14:00-16:00', 'Planning & Assessment', 'Next Week Preparation', '2 hours', 'No', '']
    ]
    
    # FIXED: Create LeetCode Progress Tracker (matching columns)
    leetcode_data = [
        ['Array Problems', 'Easy', 10, 0, '0%', 'Two Sum, Best Time to Buy/Sell'],
        ['String Problems', 'Easy', 15, 0, '0%', 'Reverse String, Valid Palindrome'],
        ['Linked List', 'Medium', 12, 0, '0%', 'Reverse List, Detect Cycle'],
        ['Tree Problems', 'Medium', 15, 0, '0%', 'BST Validation, Level Order'],
        ['Graph Problems', 'Medium', 10, 0, '0%', 'Number of Islands, Course Schedule'],
        ['Dynamic Programming', 'Hard', 8, 0, '0%', 'Knapsack, Longest Subsequence'],
        ['System Design', 'Variable', 5, 0, '0%', 'TinyURL, Twitter Clone']
    ]
    
    # Create DataFrames with CORRECT column counts
    df_main = pd.DataFrame(main_syllabus_data, 
                          columns=['Phase', 'Month', 'Week', 'Topics', 'Key Concepts', 
                                  'Projects', 'Hours', 'Status', 'Start Date', 'End Date', 'Notes'])
    
    df_progress = pd.DataFrame(progress_data,
                              columns=['Week', 'Month', 'Phase', 'Topics', 'Hours Planned',
                                      'Hours Completed', 'Completion %', 'Status', 
                                      'Project Completed', 'Notes', 'Date Completed'])
    
    df_projects = pd.DataFrame(project_data,
                              columns=['Project Name', 'Type', 'Difficulty', 'Month', 'Week',
                                      'Status', 'GitHub Link', 'Technologies Used', 
                                      'Description', 'Demo Link'])
    
    df_skills = pd.DataFrame(skills_data,
                            columns=['Skill Category', 'Specific Skill', 'Month Target',
                                    'Status', 'Confidence Level', 'Notes', 'Resources'])
    
    df_schedule = pd.DataFrame(schedule_data,
                              columns=['Day', 'Time Slot', 'Activity Type', 'Topic',
                                      'Duration', 'Completed', 'Notes'])
    
    # FIXED: Corrected LeetCode DataFrame columns
    df_leetcode = pd.DataFrame(leetcode_data,
                              columns=['Problem Category', 'Difficulty', 'Target Problems', 
                                      'Completed', 'Progress', 'Example Problems'])
    
    # Create Excel file with multiple sheets
    filename = f'DSA_Learning_Syllabus_{start_date.strftime("%Y%m%d")}.xlsx'
    
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Main Syllabus Schedule
            df_main.to_excel(writer, sheet_name='MAIN SYLLABUS SCHEDULE', index=False)
            
            # Progress Tracker
            df_progress.to_excel(writer, sheet_name='PROGRESS TRACKER', index=False)
            
            # Project Portfolio
            df_projects.to_excel(writer, sheet_name='PROJECT PORTFOLIO', index=False)
            
            # Skills Checklist
            df_skills.to_excel(writer, sheet_name='SKILLS CHECKLIST', index=False)
            
            # Weekly Schedule Template
            df_schedule.to_excel(writer, sheet_name='WEEKLY SCHEDULE', index=False)
            
            # LeetCode Tracker
            df_leetcode.to_excel(writer, sheet_name='LEETCODE TRACKER', index=False)
            
            # Adjust column widths for better readability
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"✅ Excel file '{filename}' created successfully!")
        print("📊 File contains 6 sheets:")
        print("   1. MAIN SYLLABUS SCHEDULE")
        print("   2. PROGRESS TRACKER") 
        print("   3. PROJECT PORTFOLIO")
        print("   4. SKILLS CHECKLIST")
        print("   5. WEEKLY SCHEDULE")
        print("   6. LEETCODE TRACKER")
        print(f"\n📁 File saved in your current directory")
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        return None
    
    return filename

def add_dsa_formatting(filename):
    """Add advanced formatting to the DSA Excel file"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl import load_workbook
        
        # Load the workbook
        workbook = load_workbook(filename)
        
        # Define styles for DSA phases
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        
        phase_fills = {
            'Foundation': PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
            'Linear Structures': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            'Trees & Hierarchical': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            'Graphs & Algorithms': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
            'Sorting & Searching': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            'Dynamic Programming': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            'Advanced Topics': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            'Interview Prep': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        }
        
        difficulty_fills = {
            'Beginner': PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
            'Intermediate': PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
            'Advanced': PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        }
        
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Apply formatting to all sheets
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Format header row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            # Apply phase-based coloring to MAIN SYLLABUS sheet
            if sheet_name == 'MAIN SYLLABUS SCHEDULE':
                for row in range(2, worksheet.max_row + 1):
                    phase_cell = worksheet[f'A{row}']
                    if phase_cell.value in phase_fills:
                        for col in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row, column=col).fill = phase_fills[phase_cell.value]
                            worksheet.cell(row=row, column=col).border = thin_border
            
            # Apply difficulty-based coloring to PROJECT PORTFOLIO
            if sheet_name == 'PROJECT PORTFOLIO':
                for row in range(2, worksheet.max_row + 1):
                    difficulty_cell = worksheet[f'C{row}']
                    if difficulty_cell.value in difficulty_fills:
                        difficulty_cell.fill = difficulty_fills[difficulty_cell.value]
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
            # Auto-filter on header row
            worksheet.auto_filter.ref = worksheet.dimensions
        
        # Save the formatted workbook
        formatted_filename = filename.replace('.xlsx', '_FORMATTED.xlsx')
        workbook.save(formatted_filename)
        print(f"✅ Formatted version saved as '{formatted_filename}'")
        
    except ImportError:
        print("ℹ️  For advanced formatting, install openpyxl: pip install openpyxl")
    except Exception as e:
        print(f"❌ Error formatting file: {e}")

def display_dsa_timeline():
    """Display a simple timeline of the DSA learning journey"""
    today = datetime.now()
    days_until_monday = (7 - today.weekday()) % 7
    if days_until_monday == 0:
        days_until_monday = 7
    start_date = today + timedelta(days=days_until_monday)
    
    milestones = [
        (1, "Complexity Analysis & Arrays", start_date + timedelta(weeks=3)),
        (2, "Linear Data Structures", start_date + timedelta(weeks=7)),
        (3, "Trees & Hierarchical Data", start_date + timedelta(weeks=11)),
        (4, "Graph Algorithms", start_date + timedelta(weeks=15)),
        (5, "Sorting & Searching", start_date + timedelta(weeks=19)),
        (6, "Dynamic Programming", start_date + timedelta(weeks=23)),
        (7, "Advanced Topics", start_date + timedelta(weeks=27)),
        (8, "Interview Ready", start_date + timedelta(weeks=31))
    ]
    
    print("\n📅 DSA LEARNING TIMELINE:")
    print("=" * 60)
    for milestone, description, date in milestones:
        print(f"Month {milestone}: {description:35} → {date.strftime('%b %d, %Y')}")

def display_learning_resources():
    """Display recommended learning resources"""
    print("\n📚 RECOMMENDED LEARNING RESOURCES:")
    print("=" * 50)
    resources = [
        ("Books", "Introduction to Algorithms (CLRS), Cracking the Coding Interview"),
        ("Online Platforms", "LeetCode, HackerRank, GeeksforGeeks, Educative"),
        ("YouTube Channels", "NeetCode, Back To Back SWE, Abdul Bari"),
        ("Practice Platforms", "LeetCode (300+ problems), HackerRank (DS&A track)"),
        ("Python Specific", "Python Documentation, Real Python tutorials")
    ]
    
    for category, items in resources:
        print(f"▪ {category:15}: {items}")

if __name__ == "__main__":
    print("🔬 Data Structures & Algorithms Learning Syllabus Generator")
    print("=" * 55)
    
    # Create the main Excel file
    filename = create_dsa_learning_syllabus()
    
    if filename:
        # Display timeline
        display_dsa_timeline()
        
        # Display resources
        display_learning_resources()
        
        # Ask if user wants formatted version
        try:
            format_choice = input("\nWould you like a formatted version with colors? (y/n): ").lower()
            if format_choice in ['y', 'yes']:
                add_dsa_formatting(filename)
        except:
            print("ℹ️  Basic version created successfully.")
        
        print("\n🎉 Your personalized DSA Learning Syllabus is ready!")
        print("💡 You can now open the Excel file and start your DSA journey")
        print("🚀 Happy coding! 💻")
    else:
        print("❌ Failed to create the syllabus file. Please check the error messages above.")